"""
ErgoCam v3.0 — core/session.py
Manajemen sesi: timer, pencatatan event, export CSV & XLSX.
"""

from __future__ import annotations

import csv
import os
import time
from datetime import datetime
from typing import List, Tuple

import config

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


EventRow = Tuple[str, str, str, str]   # (timestamp, elapsed, prox, slouch)


class Session:
    def __init__(self):
        self._start_time  = time.monotonic()
        self._wall_start  = datetime.now()
        self._events: List[EventRow] = []
        self._last_prox   = "idle"
        self._last_slouch = "idle"
        self._alert_count = 0

    # ── Timer ─────────────────────────────────────────────────

    @property
    def elapsed_sec(self) -> int:
        return int(time.monotonic() - self._start_time)

    def elapsed_hms(self) -> str:
        s = self.elapsed_sec
        return f"{s // 3600:02d}:{(s % 3600) // 60:02d}:{s % 60:02d}"

    def reset(self):
        self._start_time = time.monotonic()
        self._events.clear()
        self._alert_count = 0

    # ── Logging ───────────────────────────────────────────────

    def log_event(self, prox_status: str, slouch_status: str):
        """Catat event hanya bila status berubah."""
        if prox_status == self._last_prox and slouch_status == self._last_slouch:
            return
        self._last_prox   = prox_status
        self._last_slouch = slouch_status
        ts  = datetime.now().strftime("%H:%M:%S")
        ela = self.elapsed_hms()
        self._events.append((ts, ela, prox_status, slouch_status))
        if prox_status == "alert" or slouch_status == "alert":
            self._alert_count += 1

    # ── Report ────────────────────────────────────────────────

    def write_report(self):
        os.makedirs(config.REPORT_DIR, exist_ok=True)
        date_str = self._wall_start.strftime("%Y%m%d_%H%M%S")
        self._write_csv(os.path.join(config.REPORT_DIR, f"ergocam_{date_str}.csv"))
        if _HAS_OPENPYXL:
            self._write_xlsx(os.path.join(config.REPORT_DIR, f"ergocam_{date_str}.xlsx"))

    def _write_csv(self, path: str):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Timestamp", "Elapsed", "Proximity", "Posture"])
            w.writerows(self._events)

    def _write_xlsx(self, path: str):
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "ErgoCam Session"

        # Header
        headers = ["Timestamp", "Elapsed", "Proximity", "Posture"]
        hdr_fill = PatternFill("solid", fgColor="0071E3")
        hdr_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        # Data
        color_map = {
            "ok":      "34C759",
            "caution": "FF9500",
            "warn":    "FF9500",
            "alert":   "FF3B30",
            "idle":    "AEAEB2",
        }
        for row_idx, (ts, ela, prox, slouch) in enumerate(self._events, 2):
            ws.cell(row=row_idx, column=1, value=ts)
            ws.cell(row=row_idx, column=2, value=ela)
            pc = ws.cell(row=row_idx, column=3, value=prox)
            sc = ws.cell(row=row_idx, column=4, value=slouch)
            pc.fill = PatternFill("solid", fgColor=color_map.get(prox, "AEAEB2"))
            sc.fill = PatternFill("solid", fgColor=color_map.get(slouch, "AEAEB2"))

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2.append(["Session Start", self._wall_start.strftime("%Y-%m-%d %H:%M:%S")])
        ws2.append(["Session Length", self.elapsed_hms()])
        ws2.append(["Total Events",   len(self._events)])
        ws2.append(["Alert Count",    self._alert_count])

        # Column width
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16

        wb.save(path)
